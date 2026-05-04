import os
import datetime
import shutil
import sqlite3
import threading

import openpyxl
import pyzipper
import random
import string
import time
import socket
from flask import Flask, request, jsonify, send_from_directory, send_file
import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter.scrolledtext import ScrolledText
from tkinter import messagebox, filedialog
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import unicodedata
import requests

from database import init_database, get_db_connection

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

UDP_PORT = 5002

# 全局变量：控制试卷是否允许下载
PAPER_RELEASED = False
# 全局变量：远程清理触发开关
CLEAN_TRIGGER = False


def get_db():
    conn = get_db_connection()
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


# ===================== 辅助函数 =====================
def get_student_by_host(mac, hostname):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id FROM hosts WHERE mac_address=? AND hostname=?', (mac, hostname))
    host_r = c.fetchone()
    if not host_r:
        conn.close()
        return None
    host_id = host_r[0]
    c.execute('SELECT student_id FROM bindings WHERE host_id=?', (host_id,))
    bind_r = c.fetchone()
    if not bind_r:
        conn.close()
        return None
    student_id = bind_r[0]
    c.execute('SELECT id, name, student_no FROM students WHERE id=?', (student_id,))
    student = c.fetchone()
    conn.close()
    return student


def update_student_status_by_host(mac, hostname, status):
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    student = get_student_by_host(mac, hostname)
    if not student:
        return
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE students SET status=?, last_heartbeat=? WHERE id=?',
              (status, now, student[0]))
    conn.commit()
    conn.close()


# ===================== 心跳与API路由 =====================
def heartbeat_checker():
    while True:
        time.sleep(10)
        now = datetime.datetime.now()
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT id, last_heartbeat FROM students WHERE status != '未连接'")
        for sid, hb_time in c.fetchall():
            if not hb_time:
                continue
            try:
                dt = datetime.datetime.strptime(hb_time, "%Y-%m-%d %H:%M:%S")
                if (now - dt).total_seconds() > 30:
                    c.execute('UPDATE students SET status="未连接" WHERE id=?', (sid,))
            except:
                pass
        conn.commit()
        conn.close()


@app.route('/api/heartbeat', methods=['POST'])
def heartbeat():
    data = request.json
    mac = data.get("mac_address")
    host = data.get("hostname")
    ip = data.get("ip_address")
    conn = get_db()
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO hosts (mac_address, hostname, ip_address, last_seen)
                 VALUES (?, ?, ?, CURRENT_TIMESTAMP)''', (mac, host, ip))
    conn.commit()
    conn.close()
    update_student_status_by_host(mac, host, "已连接")
    return jsonify({"code": 0})


@app.route('/api/check_bind', methods=['POST'])
def check_bind():
    data = request.json
    mac = data.get('mac_address')
    host = data.get('hostname')
    student = get_student_by_host(mac, host)
    if student:
        update_student_status_by_host(mac, host, "已连接")
        return jsonify({'code': 0, 'name': student[1], 'student_no': student[2]})
    return jsonify({'code': -1})


@app.route('/api/login', methods=['POST'])
def login():
    d = request.json
    name = d['name']
    no = d['student_no']
    mac = d['mac_address']
    hostname = d['hostname']

    conn = get_db()
    c = conn.cursor()

    c.execute('SELECT id FROM students WHERE name=? AND student_no=?', (name, no))
    r = c.fetchone()
    if not r:
        conn.close()
        return jsonify({'code': -1, 'msg': '考生信息错误'})

    student_id = r[0]

    c.execute('SELECT host_id FROM bindings WHERE student_id=?', (student_id,))
    existing_bind = c.fetchone()

    c.execute('SELECT id FROM hosts WHERE mac_address=? AND hostname=?', (mac, hostname))
    current_host = c.fetchone()

    if existing_bind and current_host:
        if existing_bind[0] != current_host[0]:
            conn.close()
            return jsonify({'code': -2, 'msg': '该考生已绑定其他机器，无法在此登录'})

    if current_host:
        c.execute('SELECT student_id FROM bindings WHERE host_id=?', (current_host[0],))
        host_bind = c.fetchone()
        if host_bind and host_bind[0] != student_id:
            conn.close()
            return jsonify({'code': -3, 'msg': '该机器已绑定其他考生'})

    if current_host:
        c.execute('DELETE FROM bindings WHERE student_id=? OR host_id=?', (student_id, current_host[0]))
        c.execute('INSERT INTO bindings (student_id, host_id) VALUES (?, ?)', (student_id, current_host[0]))
        conn.commit()

    conn.close()

    update_student_status_by_host(mac, hostname, "已登录")
    return jsonify({'code': 0, 'student_id': student_id})


@app.route('/api/upload_submit', methods=['POST'])
def upload_submit():
    data = request.form
    file = request.files.get('file')
    student_id = data['student_id']
    student_no = data.get('student_no', 'unknown')

    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT name, student_no FROM students WHERE id=?', (student_id,))
    stu_info = c.fetchone()
    if not stu_info:
        conn.close()
        return jsonify({'code': -2})

    c.execute('SELECT submit_save_path_win, submit_save_path_linux FROM exams ORDER BY id DESC LIMIT 1')
    exam_config = c.fetchone()

    submit_dir = exam_config[0] if exam_config and exam_config[0] else './submissions'

    if not os.path.exists(submit_dir):
        try:
            os.makedirs(submit_dir, exist_ok=True)
        except:
            pass

    save_filename = f"{stu_info[1]}_{datetime.datetime.now().strftime('%H%M%S')}.zip"
    save_path = os.path.join(submit_dir, save_filename)

    try:
        file.save(save_path)
    except Exception as e:
        return jsonify({'code': -4, 'msg': f'保存失败: {str(e)}'})

    c.execute('''
        INSERT INTO submissions (student_id, student_name, student_no, file_path, file_size)
        VALUES (?, ?, ?, ?, ?)
    ''', (student_id, stu_info[0], stu_info[1], save_path, os.path.getsize(save_path)))
    conn.commit()
    conn.close()

    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE students SET status="已交卷" WHERE id=?', (student_id,))
    conn.commit()
    conn.close()

    return jsonify({"code": 0})


# ===================== 新API：获取考试信息（含下发状态和题量） =====================
@app.route('/api/get_exam_info', methods=['GET'])
def get_exam_info():
    global PAPER_RELEASED
    conn = get_db()
    c = conn.cursor()
    c.execute(
        'SELECT id, exam_name, exam_start_time, exam_end_time, zip_path, client_path_win, client_path_linux, question_count FROM exams ORDER BY id DESC LIMIT 1')
    exam = c.fetchone()
    conn.close()

    if not exam:
        return jsonify({"code": -1, "msg": "考试未发布"})

    return jsonify({
        "code": 0,
        "exam_id": exam[0],
        "exam_name": exam[1],
        "start_time": exam[2],
        "end_time": exam[3],
        "has_paper": True if exam[4] and os.path.exists(exam[4]) else False,
        "paper_released": PAPER_RELEASED,
        "question_count": exam[7] if exam[7] else 0,
        "client_path_win": exam[5] or "",
        "client_path_linux": exam[6] or ""
    })


# ===================== 新API：直接下载试卷文件（检查下发状态） =====================
@app.route('/api/download_paper', methods=['GET'])
def download_paper():
    global PAPER_RELEASED

    if not PAPER_RELEASED:
        return jsonify({"code": -2, "msg": "试卷尚未下发，请等待监考老师操作"}), 403

    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT zip_path FROM exams ORDER BY id DESC LIMIT 1')
    exam = c.fetchone()
    conn.close()

    if not exam or not exam[0] or not os.path.exists(exam[0]):
        return jsonify({"code": -1, "msg": "试卷未找到"}), 404

    return send_file(exam[0], as_attachment=True, download_name=os.path.basename(exam[0]))


# ===================== 修改API：考前5分钟才给密码（无限制获取） =====================
@app.route('/api/get_password', methods=['GET'])
def get_password():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT password, exam_start_time FROM exams ORDER BY id DESC LIMIT 1')
    exam = c.fetchone()
    conn.close()

    if not exam:
        return jsonify({"code": -1, "msg": "考试未发布"})

    password, start_time_str = exam

    try:
        start_time = datetime.datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
        now = datetime.datetime.now()

        time_diff = (start_time - now).total_seconds()

        if time_diff > 300:
            return jsonify({"code": -2, "msg": f"未到密码下发时间，考前5分钟开放"})

        return jsonify({"code": 0, "password": password})

    except Exception as e:
        return jsonify({"code": -3, "msg": f"时间解析错误: {str(e)}"})


@app.route('/api/get_exam_time', methods=['GET'])
def get_exam_time():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT exam_start_time, exam_end_time FROM exams ORDER BY id DESC LIMIT 1')
    r = c.fetchone()
    conn.close()
    return jsonify({"code": 0, "start": r[0], "end": r[1]} if r else {"code": -1})


# ===================== 远程清理API（静默执行版） =====================
@app.route('/api/trigger_clean', methods=['POST'])
def trigger_clean():
    """老师点击按钮：触发全局远程清理"""
    global CLEAN_TRIGGER
    CLEAN_TRIGGER = True
    return jsonify({'code': 0, 'msg': '已发送全局清理指令'})


@app.route('/api/get_clean_status', methods=['GET'])
def get_clean_status():
    """客户端轮询：是否需要执行清理"""
    global CLEAN_TRIGGER
    return jsonify({'code': 0, 'need_clean': CLEAN_TRIGGER})


@app.route('/api/clean_done', methods=['POST'])
def clean_done():
    """客户端报告：清理完成（可选，用于统计）"""
    return jsonify({'code': 0})


# ===================================================================

def run_server():
    app.run(host='0.0.0.0', port=5001, debug=False, threaded=True)


# ===================== UDP 广播 + 采集 =====================
def broadcast_scan():
    msg = b"NOI_DISCOVER"
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
    s.sendto(msg, ('<broadcast>', UDP_PORT))
    s.close()


def listen_hosts(timeout=5):
    start = time.time()
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.bind(('', 5003))
    s.settimeout(0.5)
    while time.time() - start < timeout:
        try:
            data, addr = s.recvfrom(1024)
            text = data.decode('utf-8')
            if text.startswith("NOI_HOST|"):
                _, mac, host = text.split("|")
                conn = get_db()
                c = conn.cursor()
                c.execute('''
                    INSERT OR REPLACE INTO hosts (mac_address, hostname, ip_address, last_seen)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                ''', (mac, host, addr[0]))
                conn.commit()
                conn.close()
        except:
            pass
    s.close()


def random_bind_students():
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT s.id FROM students s LEFT JOIN bindings b ON s.id = b.student_id WHERE b.student_id IS NULL")
    students = [x[0] for x in c.fetchall()]

    c.execute(
        "SELECT h.id, h.mac_address, h.hostname FROM hosts h LEFT JOIN bindings b ON h.id = b.host_id WHERE b.host_id IS NULL")
    hosts = c.fetchall()

    if len(students) == 0:
        conn.close()
        return False, "没有未绑定的考生"

    if len(hosts) == 0:
        conn.close()
        return False, "没有未绑定的主机"

    if len(students) > len(hosts):
        conn.close()
        return False, f"未绑定考生数({len(students)})大于未绑定主机数({len(hosts)})，无法完成绑定"

    random.shuffle(students)
    random.shuffle(hosts)

    for sid, host in zip(students, hosts):
        host_id, _, _ = host
        c.execute("INSERT INTO bindings (student_id, host_id) VALUES (?, ?)", (sid, host_id))

    conn.commit()
    conn.close()
    return True, f"随机绑定完成！共绑定{min(len(students), len(hosts))}名考生"


# ===================== Excel考生导入相关功能 =====================
def display_width(text):
    return sum(2 if unicodedata.east_asian_width(c) in ('F', 'W') else 1 for c in str(text or ''))


def auto_fit_columns(ws, min_w=8, max_w=50, padding=3):
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        w = max((display_width(c.value) for c in col_cells
                 if not isinstance(c, openpyxl.cell.cell.MergedCell) and c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(w * 1.1 + padding, max_w))


def generate_student_import_template():
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "考生数据"
    header_fill = PatternFill('solid', fgColor='0070C0')
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    headers = ['姓名', '考号']
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    example_data = [['张三', '2025001'], ['李四', '2025002']]
    for row, data in enumerate(example_data, 2):
        for col, val in enumerate(data, 1):
            cell = ws_data.cell(row=row, column=col, value=val)
            cell.alignment = left_align
            cell.font = Font(name='微软雅黑', size=10)
    auto_fit_columns(ws_data)
    ws_data.freeze_panes = 'A2'
    ws_note = wb.create_sheet("填写说明")
    notes = ['填写说明：', '1. 姓名、考号为必填项', '2. 考号不能重复', '3. 无需填写MAC，导入后通过绑定功能分配机器']
    for i, note in enumerate(notes, 1):
        cell = ws_note.cell(row=i, column=1, value=note)
        cell.font = Font(name='微软雅黑', size=11)
        if i == 1: cell.font = Font(name='微软雅黑', bold=True, size=12, color='FF0000')
    template_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'NOI考生导入模板.xlsx')
    wb.save(template_path)
    return template_path


def import_students_from_excel(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name="考生数据", header=0)
        df.columns = df.columns.str.strip()
        if '姓名' not in df.columns or '考号' not in df.columns:
            return False, "导入失败：请使用系统生成的标准模板"
        df['姓名'] = df['姓名'].astype(str).str.strip()
        df['考号'] = df['考号'].astype(str).str.strip()
        df = df[(df['姓名'] != 'nan') & (df['姓名'] != '') & (df['考号'] != 'nan') & (df['考号'] != '')]
        if df.empty: return False, "导入失败：未读取到有效考生数据"
        duplicate_sno = df[df.duplicated('考号', keep=False)]
        if len(duplicate_sno) > 0: return False, f"导入失败：存在重复考号【{list(duplicate_sno['考号'].unique())}】"

        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT student_no FROM students')
        existing_snos = {x[0] for x in c.fetchall()}
        duplicate_in_db = df[df['考号'].isin(existing_snos)]
        if len(duplicate_in_db) > 0:
            conn.close()
            return False, f"导入失败：考号【{list(duplicate_in_db['考号'].unique())}】已存在"

        insert_data = [(row['姓名'], row['考号']) for _, row in df.iterrows()]
        c.executemany('INSERT INTO students (name, student_no) VALUES (?, ?)', insert_data)
        conn.commit()
        conn.close()
        return True, f"导入成功！共导入{len(df)}名考生"
    except Exception as e:
        return False, f"导入失败：{str(e)}"


# ===================== GUI =====================
class ServerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title('NOI 试题收发系统 - 教师控制台')
        self.root.geometry('1600x1200')
        self.root.place_window_center()

        header = ttk.Frame(root, padding=10)
        header.pack(fill=X)
        ttk.Label(header, text='NOI 服务器管理中心', font=('Helvetica', 18, 'bold'), bootstyle=PRIMARY).pack(side=LEFT)

        main_frame = ttk.Frame(root, padding=10)
        main_frame.pack(fill=BOTH, expand=True)

        left_panel = ttk.Frame(main_frame)
        left_panel.pack(side=LEFT, fill=Y, expand=False, padx=(0, 10))

        # ===== 考生管理 =====
        group_student = ttk.Labelframe(left_panel, text=' 👨‍🎓 考生管理 ', padding=15)
        group_student.pack(fill=X, pady=(0, 15))

        ttk.Label(group_student, text='姓名').grid(row=0, column=0, pady=5, sticky=W)
        self.name = ttk.Entry(group_student, width=15)
        self.name.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(group_student, text='考号').grid(row=0, column=2, pady=5, sticky=W)
        self.sno = ttk.Entry(group_student, width=15)
        self.sno.grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(group_student, text='绑定主机').grid(row=1, column=0, pady=5, sticky=W)
        self.host_combobox = ttk.Combobox(group_student, width=32, state='readonly')
        self.host_combobox.grid(row=1, column=1, columnspan=3, padx=5, pady=5)
        self.host_combobox.set("暂不绑定 / 稍后随机分配")

        ttk.Button(group_student, text='➕ 添加/绑定考生', bootstyle=SUCCESS, command=self.add_student).grid(row=2,
                                                                                                            column=0,
                                                                                                            columnspan=4,
                                                                                                            pady=10,
                                                                                                            sticky=EW)
        ttk.Button(group_student, text='🔄 刷新主机列表', bootstyle=(SECONDARY, OUTLINE),
                   command=self.refresh_host_list).grid(row=3, column=0, columnspan=4, sticky=EW)

        # Excel导入
        group_import = ttk.Labelframe(left_panel, text=' 📊 Excel批量导入 ', padding=15)
        group_import.pack(fill=X, pady=(0, 15))
        ttk.Button(group_import, text='📋 下载模板', bootstyle=INFO, command=self.download_import_template).grid(row=0,
                                                                                                                column=0,
                                                                                                                padx=5,
                                                                                                                pady=5,
                                                                                                                sticky=EW)
        ttk.Button(group_import, text='📁 导入Excel', bootstyle=PRIMARY, command=self.import_students_excel).grid(row=0,
                                                                                                                 column=1,
                                                                                                                 padx=5,
                                                                                                                 pady=5,
                                                                                                                 sticky=EW)
        group_import.grid_columnconfigure(0, weight=1)
        group_import.grid_columnconfigure(1, weight=1)

        # ===== 考试配置 =====
        group_exam = ttk.Labelframe(left_panel, text=' 📝 考试配置 ', padding=15)
        group_exam.pack(fill=X, pady=(0, 15))

        ttk.Label(group_exam, text='考试名称').grid(row=0, column=0, pady=5, sticky=W)
        self.ename = ttk.Entry(group_exam)
        self.ename.grid(row=0, column=1, columnspan=2, sticky=EW, padx=5, pady=5)

        ttk.Label(group_exam, text='解压密码').grid(row=1, column=0, pady=5, sticky=W)
        self.pwd = ttk.Entry(group_exam, width=12)
        self.pwd.grid(row=1, column=1, sticky=W, padx=5, pady=5)
        ttk.Button(group_exam, text='随机', bootstyle=(INFO, OUTLINE), command=self.gen_rand_pwd).grid(row=1, column=2,
                                                                                                       padx=5)

        ttk.Label(group_exam, text='试题数量').grid(row=2, column=0, pady=5, sticky=W)
        self.question_count = ttk.Entry(group_exam, width=8)
        self.question_count.grid(row=2, column=1, sticky=W, padx=5, pady=5)

        ttk.Label(group_exam, text='原题目录').grid(row=3, column=0, pady=5, sticky=W)
        self.zip_path_entry = ttk.Entry(group_exam)
        self.zip_path_entry.grid(row=3, column=1, sticky=EW, padx=5, pady=5)
        ttk.Button(group_exam, text='浏览', bootstyle=(SECONDARY, OUTLINE), command=self.browse_zip).grid(row=3,
                                                                                                          column=2,
                                                                                                          padx=5)

        ttk.Label(group_exam, text='客户端Win路径').grid(row=4, column=0, pady=5, sticky=W)
        self.client_path_win = ttk.Entry(group_exam)
        self.client_path_win.insert(0, "D:\\NOI_Exam\\Paper")
        self.client_path_win.grid(row=4, column=1, columnspan=2, sticky=EW, padx=5, pady=5)

        ttk.Label(group_exam, text='客户端Linux路径').grid(row=5, column=0, pady=5, sticky=W)
        self.client_path_linux = ttk.Entry(group_exam)
        self.client_path_linux.insert(0, "/home/NOI_Exam/Paper")
        self.client_path_linux.grid(row=5, column=1, columnspan=2, sticky=EW, padx=5, pady=5)

        ttk.Label(group_exam, text='服务端收卷路径').grid(row=6, column=0, pady=5, sticky=W)
        self.submit_save_path = ttk.Entry(group_exam)
        self.submit_save_path.insert(0, os.path.join(os.path.expanduser('~'), 'Desktop', 'NOI_Submissions'))
        self.submit_save_path.grid(row=6, column=1, sticky=EW, padx=5, pady=5)
        ttk.Button(group_exam, text='浏览', bootstyle=(SECONDARY, OUTLINE), command=self.select_submit_path).grid(row=6,
                                                                                                                  column=2,
                                                                                                                  padx=5)

        ttk.Label(group_exam, text='开始时间').grid(row=7, column=0, pady=5, sticky=W)
        self.etime_start = ttk.Entry(group_exam)
        self.etime_start.insert(0, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.etime_start.grid(row=7, column=1, columnspan=2, sticky=EW, padx=5, pady=5)

        ttk.Label(group_exam, text='结束时间').grid(row=8, column=0, pady=5, sticky=W)
        self.etime_end = ttk.Entry(group_exam)
        self.etime_end.insert(0, (datetime.datetime.now() + datetime.timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S"))
        self.etime_end.grid(row=8, column=1, columnspan=2, sticky=EW, padx=5, pady=5)

        ttk.Button(
            group_exam,
            text='🚀 发布考试',
            bootstyle=PRIMARY,
            command=self.start_set_exam
        ).grid(row=9, column=0, columnspan=3, pady=15, sticky=EW)

        right_panel = ttk.Frame(main_frame)
        right_panel.pack(side=LEFT, fill=BOTH, expand=True)

        group_ctrl = ttk.Labelframe(right_panel, text=' ⚡ 考场控制面板 ', padding=15)
        group_ctrl.pack(fill=X, pady=(0, 15))

        btn_frame = ttk.Frame(group_ctrl)
        btn_frame.pack(fill=X)
        ttk.Button(btn_frame, text='📡 采集局域网主机', bootstyle=WARNING, command=self.collect_hosts).pack(side=LEFT,
                                                                                                           padx=5,
                                                                                                           expand=True,
                                                                                                           fill=X)
        ttk.Button(btn_frame, text='🎲 随机绑定考生', bootstyle=INFO, command=self.random_bind).pack(side=LEFT, padx=5,
                                                                                                    expand=True, fill=X)
        ttk.Button(btn_frame, text='🖥️ 查看主机', bootstyle=SECONDARY, command=self.show_hosts).pack(side=LEFT, padx=5,
                                                                                                     expand=True,
                                                                                                     fill=X)
        ttk.Button(btn_frame, text='📋 查看考生', bootstyle=SUCCESS, command=self.show_all_students).pack(side=LEFT,
                                                                                                         padx=5,
                                                                                                         expand=True,
                                                                                                         fill=X)
        ttk.Button(btn_frame, text='📂 查看提交', bootstyle=INFO, command=self.show_submits).pack(side=LEFT, padx=5,
                                                                                                 expand=True, fill=X)

        btn_frame2 = ttk.Frame(group_ctrl)
        btn_frame2.pack(fill=X, pady=(5, 0))
        ttk.Button(btn_frame2, text='📤 手动下发试卷', bootstyle=(SUCCESS, OUTLINE), command=self.release_paper).pack(
            side=LEFT, padx=5, expand=True, fill=X)
        ttk.Button(btn_frame2, text='🔓 解除所有绑定', bootstyle=(DANGER, OUTLINE), command=self.unbind_all).pack(
            side=LEFT, padx=5, expand=True, fill=X)
        ttk.Button(btn_frame2, text='🗑️ 一键清空考场', bootstyle=DANGER, command=self.clear_all_data).pack(side=LEFT,
                                                                                                           padx=5,
                                                                                                           expand=True,
                                                                                                           fill=X)

        # ===== 新增：远程清理按钮 =====
        btn_frame3 = ttk.Frame(group_ctrl)
        btn_frame3.pack(fill=X, pady=(5, 0))
        ttk.Button(btn_frame3, text='🧹 远程清空所有考生目录', bootstyle=DANGER, command=self.remote_clean_all).pack(
            side=LEFT, padx=5, expand=True, fill=X)

        group_log = ttk.Labelframe(right_panel, text=' 📟 系统运行日志 ', padding=10)
        group_log.pack(fill=BOTH, expand=True)
        self.log = ScrolledText(group_log, wrap=WORD)
        self.log.pack(fill=BOTH, expand=True)

        self.log_print('✅ 服务端已启动 (标准化考试模式)')
        self.refresh_host_list()

    def log_print(self, msg):
        self.log.insert(END, f'[{datetime.datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self.log.see(END)

    def refresh_host_list(self):
        conn = get_db()
        c = conn.cursor()
        c.execute('''
            SELECT h.id, h.hostname, h.mac_address 
            FROM hosts h 
            LEFT JOIN bindings b ON h.id = b.host_id 
            WHERE b.host_id IS NULL
            ORDER BY h.last_seen DESC
        ''')
        hosts = c.fetchall()
        conn.close()

        values = ["暂不绑定 / 稍后随机分配"]
        self.host_map = {}

        for h_id, h_name, h_mac in hosts:
            display_text = f"{h_name} ({h_mac})"
            values.append(display_text)
            self.host_map[display_text] = h_id

        self.host_combobox['values'] = values
        if not self.host_combobox.get():
            self.host_combobox.set(values[0])
        self.log_print(f"🔄 主机列表已刷新，共{len(hosts)}台未绑定主机")

    def collect_hosts(self):
        self.log_print("🧹 正在清理历史采集数据...")
        try:
            conn = get_db()
            c = conn.cursor()
            c.execute("DELETE FROM hosts")
            conn.commit()
            conn.close()
        except Exception as e:
            self.log_print(f"❌ 清理失败: {e}")

        self.log_print("📡 开始广播采集主机...")
        broadcast_scan()
        listen_hosts(timeout=5)
        self.log_print("✅ 主机采集完成")
        self.refresh_host_list()
        messagebox.showinfo("完成", "主机采集完成")

    def random_bind(self):
        ok, msg = random_bind_students()
        if ok:
            self.log_print("✅ " + msg)
            self.refresh_host_list()
            messagebox.showinfo("完成", msg)
        else:
            self.log_print("❌ " + msg)
            messagebox.showerror("失败", msg)

    def unbind_all(self):
        if messagebox.askyesno("确认", "确定要解除所有考生与主机的绑定关系吗？"):
            conn = get_db()
            c = conn.cursor()
            c.execute("DELETE FROM bindings")
            conn.commit()
            conn.close()
            self.log_print("🔓 所有绑定关系已解除")
            self.refresh_host_list()
            messagebox.showinfo("完成", "所有绑定关系已解除")

    def clear_all_data(self):
        global PAPER_RELEASED
        if messagebox.askyesno("⚠️ 危险操作",
                               "确定要清空所有数据吗？\n\n这将删除：\n- 所有考生信息\n- 所有主机信息\n- 所有绑定关系\n- 所有提交记录\n\n此操作不可恢复！",
                               icon='warning'):
            try:
                PAPER_RELEASED = False
                conn = get_db()
                c = conn.cursor()
                c.execute("DELETE FROM submissions")
                c.execute("DELETE FROM bindings")
                c.execute("DELETE FROM students")
                c.execute("DELETE FROM hosts")
                c.execute("DELETE FROM clean_commands")
                conn.commit()
                conn.close()
                self.log_print("🗑️ 考场数据已全部清空，可以开始下一场考试")
                self.refresh_host_list()
                messagebox.showinfo("完成", "考场数据已清空！\n可以导入新的考生信息开始下一场考试。")
            except Exception as e:
                self.log_print(f"❌ 清空失败：{e}")
                messagebox.showerror("错误", f"清空失败：{str(e)}")

    def release_paper(self):
        '''手动下发试卷'''
        global PAPER_RELEASED
        if messagebox.askyesno("确认", "确定要向所有考生下发试卷吗？\n\n下发后，考生即可点击“获取试卷”按钮下载。"):
            PAPER_RELEASED = True
            self.log_print("📤 试卷已全局下发！考生现在可以下载了。")
            messagebox.showinfo("成功", "试卷已下发！\n考生端现在可以点击“获取试卷”进行下载。")

    def select_submit_path(self):
        '''选择服务端收卷目录'''
        path = filedialog.askdirectory(title='选择服务端收卷目录')
        if path:
            self.submit_save_path.delete(0, END)
            self.submit_save_path.insert(0, path)
            self.log_print(f'📂 收卷目录已设置为: {path}')

    def browse_zip(self):
        path = filedialog.askdirectory(title='选择原题文件夹')
        if path:
            self.zip_path_entry.delete(0, END)
            self.zip_path_entry.insert(0, path)

    def gen_rand_pwd(self):
        pwd = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(8))
        self.pwd.delete(0, END)
        self.pwd.insert(0, pwd)
        self.log_print(f'🔑 生成随机密码：{pwd}')

    def add_student(self):
        if not self.name.get() or not self.sno.get():
            messagebox.showwarning("警告", "姓名和考号不能为空！")
            return
        try:
            conn = get_db()
            c = conn.cursor()

            c.execute('INSERT INTO students (name,student_no) VALUES (?,?)',
                      (self.name.get(), self.sno.get()))
            student_id = c.lastrowid

            selected_text = self.host_combobox.get()
            if selected_text in self.host_map:
                host_id = self.host_map[selected_text]

                c.execute('SELECT student_id FROM bindings WHERE host_id=?', (host_id,))
                if c.fetchone():
                    conn.rollback()
                    conn.close()
                    messagebox.showerror("错误", "该主机已被其他考生绑定！")
                    return

                c.execute('INSERT INTO bindings (student_id, host_id) VALUES (?, ?)', (student_id, host_id))
                self.log_print(f'👤 考生添加成功，并已绑定至: {selected_text}')
            else:
                self.log_print(f'👤 考生添加成功：{self.name.get()} ({self.sno.get()})')

            conn.commit()
            conn.close()

            self.name.delete(0, END)
            self.sno.delete(0, END)
            self.refresh_host_list()

        except Exception as e:
            self.log_print(f'❌ 添加失败：{e}')
            messagebox.showerror("错误", f"添加失败：{str(e)}")

    def start_set_exam(self):
        global PAPER_RELEASED
        exam_name = self.ename.get().strip()
        source_path = self.zip_path_entry.get().strip()
        pwd = self.pwd.get().strip()
        etime_start = self.etime_start.get().strip()
        etime_end = self.etime_end.get().strip()

        client_path_win = self.client_path_win.get().strip()
        client_path_linux = self.client_path_linux.get().strip()
        submit_path = self.submit_save_path.get().strip()

        if not source_path or not os.path.isdir(source_path):
            return messagebox.showerror("错误", "请选择有效的试题文件夹")
        if not pwd:
            return messagebox.showerror("错误", "请设置解压密码")
        if not submit_path:
            return messagebox.showerror("错误", "请选择服务端收卷目录")

        try:
            n = int(self.question_count.get().strip())
            if n <= 0:
                return messagebox.showerror("错误", "试题数量必须大于0")
        except ValueError:
            return messagebox.showerror("错误", "试题数量必须是数字")

        PAPER_RELEASED = False
        threading.Thread(target=self.set_exam_worker, args=(
        exam_name, source_path, submit_path, client_path_win, client_path_linux, pwd, etime_start, etime_end, n),
                         daemon=True).start()

    def set_exam_worker(self, exam_name, source_path, submit_path, client_path_win, client_path_linux, pwd, etime_start,
                        etime_end, n):
        temp_dir = None
        try:
            os.makedirs('exams', exist_ok=True)

            # ===================== 核心修改：自动清理旧试卷压缩包 =====================
            self.log_print("🧹 开始自动清理旧试卷压缩包...")
            exams_dir = 'exams'
            cleaned_count = 0
            # 遍历目录，删除所有旧的zip试卷文件
            for file_name in os.listdir(exams_dir):
                file_path = os.path.join(exams_dir, file_name)
                # 仅清理ZIP格式的试卷文件，避免误删其他文件
                if os.path.isfile(file_path) and file_name.lower().endswith('.zip'):
                    try:
                        os.remove(file_path)
                        cleaned_count += 1
                        self.log_print(f"   ✅ 已删除旧试卷：{file_name}")
                    except Exception as e:
                        self.log_print(f"   ❌ 删除旧试卷失败 {file_name}：{str(e)}")
            self.log_print(f"✅ 旧试卷清理完成，共清理 {cleaned_count} 个历史文件")
            # ========================================================================

            # 生成新试卷文件名（带时间戳，唯一不重复）
            dest_filename = f"{exam_name}_{int(datetime.datetime.now().timestamp())}.zip"
            dest_path = os.path.join('exams', dest_filename)
            temp_dir = f"temp_{int(datetime.datetime.now().timestamp())}"
            os.makedirs(temp_dir, exist_ok=True)

            # 复制原题文件到临时目录
            for item in os.listdir(source_path):
                s = os.path.join(source_path, item)
                d = os.path.join(temp_dir, item)
                if os.path.isdir(s):
                    shutil.copytree(s, d, dirs_exist_ok=True)
                else:
                    shutil.copy(s, d)

            # 强制生成精确数量的答题文件夹
            # 1. 先删除所有已有的答题文件夹，避免数量混乱
            for item in os.listdir(temp_dir):
                item_path = os.path.join(temp_dir, item)
                if os.path.isdir(item_path) and item.startswith("答题文件夹_"):
                    shutil.rmtree(item_path)
            self.log_print(f'🧹 已清理临时目录中所有旧的答题文件夹')

            # 2. 精确生成从1到n的答题文件夹
            for i in range(1, n + 1):
                f = os.path.join(temp_dir, f"答题文件夹_{i}")
                os.makedirs(f, exist_ok=True)
                # 创建占位文件，确保空文件夹被打包
                open(os.path.join(f, ".keep"), 'a').close()

            # 加密压缩生成新试卷
            with pyzipper.AESZipFile(dest_path, 'w', compression=pyzipper.ZIP_DEFLATED,
                                     encryption=pyzipper.WZ_AES) as zf:
                zf.setpassword(pwd.encode())
                for root_, _, files in os.walk(temp_dir):
                    for file in files:
                        fp = os.path.join(root_, file)
                        zf.write(fp, os.path.relpath(fp, temp_dir))

            # 保存考试信息到数据库
            conn = get_db()
            c = conn.cursor()
            c.execute(
                'INSERT INTO exams (exam_name, zip_path, password, exam_start_time, exam_end_time, question_count, submit_save_path_win, submit_save_path_linux, client_path_win, client_path_linux) VALUES (?,?,?,?,?,?,?,?,?,?)',
                (exam_name, dest_path, pwd, etime_start, etime_end, n, submit_path, submit_path, client_path_win,
                 client_path_linux))

            conn.commit()
            conn.close()
            self.log_print('🎉 新考试发布成功！')
            self.log_print(f'   -> 试题数量: {n} 题（已存入数据库）')
            self.log_print(f'   -> 已精确生成 {n} 个答题文件夹（答题文件夹_1 至 答题文件夹_{n}）')
            self.log_print(f'   -> Windows客户端路径: {client_path_win}')
            self.log_print(f'   -> Linux客户端路径: {client_path_linux}')
            self.log_print(f'   -> 服务端收卷路径: {submit_path}')
            self.log_print(f'   -> 请点击【手动下发试卷】按钮允许考生下载')
            messagebox.showinfo("成功",
                                f"考试发布完成！\n\n✅ 已精确生成 {n} 个答题文件夹\n\n⚠️ 重要提示：\n请点击【手动下发试卷】按钮，考生才能下载试卷。")
        except Exception as e:
            self.log_print(f'❌ 发布失败：{e}')
        finally:
            # 清理临时文件夹
            if temp_dir and os.path.isdir(temp_dir):
                shutil.rmtree(temp_dir)

    def remote_clean_all(self):
        """一键远程清空所有考生工作目录"""
        global CLEAN_TRIGGER
        if messagebox.askyesno("⚠️ 强制操作",
                               "确定要远程清空所有已连接考生的工作目录吗？\n\n⚠️ 此操作将**强制删除**考生电脑上工作目录内的所有文件！\n⚠️ 考生端将**静默执行**，无任何弹窗提示！\n\n此操作不可恢复！",
                               icon='warning'):
            try:
                CLEAN_TRIGGER = True
                self.log_print("🧹 已向所有在线客户端发送全局强制清理指令")
                self.log_print("   -> 客户端将在3秒内开始静默执行清理")
                messagebox.showinfo("指令已发送", "全局强制清理指令已下发！\n\n所有在线考生的工作目录将被自动清空。")
            except Exception as e:
                self.log_print(f"❌ 发送指令失败: {e}")

    def show_submits(self):
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT student_no,student_name,upload_time FROM submissions ORDER BY upload_time DESC')
        data = c.fetchall()
        conn.close()
        msg = '\n'.join([f'✅ {x[0]} ({x[1]}) - {x[2]}' for x in data]) if data else '暂无提交记录'
        messagebox.showinfo('📦 收卷记录', msg)

    def show_hosts(self):
        win = tk.Toplevel(self.root)
        win.title("🖥️ 采集到的主机")
        win.geometry("750x400")
        win.transient(self.root)

        # 手动计算屏幕居中位置
        win.update_idletasks()
        screen_width = win.winfo_screenwidth()
        screen_height = win.winfo_screenheight()
        window_width = win.winfo_width()
        window_height = win.winfo_height()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        win.geometry(f"+{x}+{y}")

        tree = ttk.Treeview(win, columns=("id", "mac", "hostname", "ip", "last_seen", "status"), show="headings")
        tree.heading("id", text="ID")
        tree.heading("mac", text="MAC地址")
        tree.heading("hostname", text="主机名")
        tree.heading("ip", text="IP地址")
        tree.heading("last_seen", text="采集时间")
        tree.heading("status", text="绑定状态")

        tree.column("id", width=50, anchor=CENTER)
        tree.column("mac", width=150)
        tree.column("hostname", width=150)
        tree.column("ip", width=120)
        tree.column("last_seen", width=150)
        tree.column("status", width=80, anchor=CENTER)

        tree.pack(fill=BOTH, expand=True, padx=10, pady=10)

        conn = get_db()
        c = conn.cursor()
        c.execute('''
            SELECT h.id, h.mac_address, h.hostname, h.ip_address, h.last_seen,
                   CASE WHEN b.host_id IS NOT NULL THEN '已绑定' ELSE '未绑定' END
            FROM hosts h
            LEFT JOIN bindings b ON h.id = b.host_id
            ORDER BY h.last_seen DESC
        ''')
        # 修复：将sqlite3.Row对象转换为普通元组
        for row in c.fetchall():
            tree.insert("", END, values=tuple(row))
        conn.close()

    def show_all_students(self):
        win = tk.Toplevel(self.root)
        win.title("📋 考生信息与绑定状态")
        win.geometry("1250x500")
        win.transient(self.root)

        # 手动计算屏幕居中位置
        win.update_idletasks()
        screen_width = win.winfo_screenwidth()
        screen_height = win.winfo_screenheight()
        window_width = win.winfo_width()
        window_height = win.winfo_height()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        win.geometry(f"+{x}+{y}")

        frame = ttk.Frame(win)
        frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        columns = ("id", "name", "student_no", "status", "bind_mac", "bind_host", "bind_time")
        tree = ttk.Treeview(frame, columns=columns, show="headings")

        tree.heading("id", text="ID")
        tree.heading("name", text="姓名")
        tree.heading("student_no", text="考号")
        tree.heading("status", text="状态")
        tree.heading("bind_mac", text="绑定机器MAC")
        tree.heading("bind_host", text="绑定机器名")
        tree.heading("bind_time", text="导入时间")

        tree.column("id", width=40, anchor=CENTER)
        tree.column("name", width=80)
        tree.column("student_no", width=110)
        tree.column("status", width=100, anchor=CENTER)
        tree.column("bind_mac", width=160)
        tree.column("bind_host", width=120)

        tree.pack(fill=BOTH, expand=True)

        context_menu = tk.Menu(tree, tearoff=0)
        context_menu.add_command(label="🔓 解除绑定", command=lambda: self.unbind_selected(tree))
        context_menu.add_separator()
        context_menu.add_command(label="🔄 更换绑定主机", command=lambda: self.change_host_selected(tree, win))

        def show_context_menu(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                context_menu.post(event.x_root, event.y_root)

        tree.bind("<Button-3>", show_context_menu)

        def fmt_status(s):
            if s == "未连接":
                return "🔴 未连接"
            elif s == "已连接":
                return "🟡 已连接"
            elif s == "已登录":
                return "🟢 已登录"
            elif s == "已交卷":
                return "🔵 已交卷"
            return s

        def load_data():
            for i in tree.get_children(): tree.delete(i)
            conn = get_db()
            c = conn.cursor()
            c.execute('''
                SELECT s.id, s.name, s.student_no, s.status, 
                       h.mac_address, h.hostname, s.bind_time
                FROM students s
                LEFT JOIN bindings b ON s.id = b.student_id
                LEFT JOIN hosts h ON b.host_id = h.id
                ORDER BY s.id ASC
            ''')
            # 修复：将sqlite3.Row对象转换为普通元组
            for row in c.fetchall():
                lst = list(tuple(row))
                lst[3] = fmt_status(lst[3])
                if not lst[4]: lst[4] = "未绑定"
                if not lst[5]: lst[5] = "未绑定"
                tree.insert("", END, values=lst)
            conn.close()

        load_data()

        def refresh():
            if win.winfo_exists():
                load_data()
                win.after(5000, refresh)

        refresh()

        self.student_tree = tree
        self.student_win = win
        self.fmt_status_func = fmt_status

    def unbind_selected(self, tree):
        selected = tree.selection()
        if not selected:
            return

        item = tree.item(selected[0])
        student_id = item['values'][0]
        student_name = item['values'][1]
        bind_mac = item['values'][4]

        if bind_mac == "未绑定":
            messagebox.showinfo("提示", "该考生未绑定任何机器")
            return

        if messagebox.askyesno("确认", f"确定要解除考生【{student_name}】的绑定吗？"):
            conn = get_db()
            c = conn.cursor()
            c.execute('DELETE FROM bindings WHERE student_id=?', (student_id,))
            conn.commit()
            conn.close()
            self.log_print(f"🔓 已解除考生【{student_name}】的绑定")
            self.refresh_host_list()
            for i in tree.get_children(): tree.delete(i)
            conn = get_db()
            c = conn.cursor()
            c.execute('''
                SELECT s.id, s.name, s.student_no, s.status, 
                       h.mac_address, h.hostname, s.bind_time
                FROM students s
                LEFT JOIN bindings b ON s.id = b.student_id
                LEFT JOIN hosts h ON b.host_id = h.id
                ORDER BY s.id ASC
            ''')
            # 修复：将sqlite3.Row对象转换为普通元组
            for row in c.fetchall():
                lst = list(tuple(row))
                lst[3] = self.fmt_status_func(lst[3])
                if not lst[4]: lst[4] = "未绑定"
                if not lst[5]: lst[5] = "未绑定"
                tree.insert("", END, values=lst)
            conn.close()

    def change_host_selected(self, tree, parent_win):
        selected = tree.selection()
        if not selected:
            return

        item = tree.item(selected[0])
        student_id = item['values'][0]
        student_name = item['values'][1]

        conn = get_db()
        c = conn.cursor()
        c.execute('''
            SELECT h.id, h.hostname, h.mac_address 
            FROM hosts h 
            LEFT JOIN bindings b ON h.id = b.host_id 
            WHERE b.host_id IS NULL
            ORDER BY h.last_seen DESC
        ''')
        available_hosts = c.fetchall()
        conn.close()

        if not available_hosts:
            messagebox.showinfo("提示", "没有可用的未绑定主机！")
            return

        select_win = tk.Toplevel(parent_win)
        select_win.title(f"为【{student_name}】更换主机")
        select_win.geometry("500x300")
        select_win.transient(parent_win)

        # 手动计算屏幕居中位置
        select_win.update_idletasks()
        sw = select_win.winfo_screenwidth()
        sh = select_win.winfo_screenheight()
        ww = select_win.winfo_width()
        wh = select_win.winfo_height()
        x = (sw - ww) // 2
        y = (sh - wh) // 2
        select_win.geometry(f"+{x}+{y}")

        ttk.Label(select_win, text=f"请为考生【{student_name}】选择一台新主机：", padding=10).pack(fill=X)

        host_listbox = tk.Listbox(select_win, font=('微软雅黑', 10))
        host_listbox.pack(fill=BOTH, expand=True, padx=10, pady=5)

        host_map = {}
        for h_id, h_name, h_mac in available_hosts:
            display_text = f"{h_name} ({h_mac})"
            host_listbox.insert(END, display_text)
            host_map[display_text] = h_id

        def confirm_change():
            selection = host_listbox.curselection()
            if not selection:
                messagebox.showwarning("提示", "请选择一台主机")
                return

            selected_text = host_listbox.get(selection[0])
            new_host_id = host_map[selected_text]

            if messagebox.askyesno("确认", f"确定将【{student_name}】绑定到\n{selected_text}？"):
                conn = get_db()
                c = conn.cursor()
                c.execute('DELETE FROM bindings WHERE student_id=?', (student_id,))
                c.execute('INSERT INTO bindings (student_id, host_id) VALUES (?, ?)', (student_id, new_host_id))
                conn.commit()
                conn.close()

                self.log_print(f"🔄 已将【{student_name}】更换至 {selected_text}")
                self.refresh_host_list()
                select_win.destroy()

                for i in tree.get_children(): tree.delete(i)
                conn = get_db()
                c = conn.cursor()
                c.execute('''
                    SELECT s.id, s.name, s.student_no, s.status, 
                           h.mac_address, h.hostname, s.bind_time
                    FROM students s
                    LEFT JOIN bindings b ON s.id = b.student_id
                    LEFT JOIN hosts h ON b.host_id = h.id
                    ORDER BY s.id ASC
                ''')
                # 修复：将sqlite3.Row对象转换为普通元组
                for row in c.fetchall():
                    lst = list(tuple(row))
                    lst[3] = self.fmt_status_func(lst[3])
                    if not lst[4]: lst[4] = "未绑定"
                    if not lst[5]: lst[5] = "未绑定"
                    tree.insert("", END, values=lst)
                conn.close()

        ttk.Button(select_win, text="确认更换", bootstyle=SUCCESS, command=confirm_change).pack(pady=10)

    def fmt_status_func(self, s):
        if s == "未连接":
            return "🔴 未连接"
        elif s == "已连接":
            return "🟡 已连接"
        elif s == "已登录":
            return "🟢 已登录"
        elif s == "已交卷":
            return "🔵 已交卷"
        return s

    def download_import_template(self):
        try:
            template_path = generate_student_import_template()
            self.log_print(f'📋 模板已生成：{template_path}')
            messagebox.showinfo("成功", f"模板已生成：\n{template_path}")
        except Exception as e:
            self.log_print(f'❌ 生成模板失败：{e}')
            messagebox.showerror("错误", f"生成模板失败：{str(e)}")

    def import_students_excel(self):
        file_path = filedialog.askopenfilename(
            title='选择考生Excel文件',
            filetypes=[('Excel文件', '*.xlsx;*.xls')],
            initialdir=os.path.expanduser('~')
        )
        if not file_path: return
        self.log_print(f'📁 正在导入：{file_path}')
        ok, msg = import_students_from_excel(file_path)
        if ok:
            self.log_print(f'✅ {msg}')
            messagebox.showinfo("成功", msg)
        else:
            self.log_print(f'❌ {msg}')
            messagebox.showerror("错误", msg)


if __name__ == '__main__':
    # init_database()
    # threading.Thread(target=heartbeat_checker, daemon=True).start()
    # threading.Thread(target=run_server, daemon=True).start()
    # app_window = ttk.Window(themename="flatly")
    # ServerGUI(app_window)
    # app_window.mainloop()
    init_database()
    
    # 1. 先创建UI窗口，并将主题改为和 client.py 一样的 litera (flatly在部分mac上也有bug)
    app_window = ttk.Window(themename="litera") 
    
    # 2. 封装后台线程
    def start_background_tasks():
        threading.Thread(target=heartbeat_checker, daemon=True).start()
        threading.Thread(target=run_server, daemon=True).start()
        
    # 3. 让 GUI 渲染 100 毫秒后，再启动后台网络和检测服务
    app_window.after(100, start_background_tasks)
    
    ServerGUI(app_window)
    
    # 4. 强制刷新一次 macOS 渲染缓冲
    app_window.update_idletasks() 
    
    app_window.mainloop()